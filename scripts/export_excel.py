"""
Export benchmark result JSON to Excel report.

Usage:
    python export_excel.py results/0040_mindgpt-mindgpt-pro-v3-2-agent.json
    python export_excel.py results/0040_mindgpt-mindgpt-pro-v3-2-agent.json -o report.xlsx
    python export_excel.py results/  # process all JSON files in directory
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

PASS_THRESHOLD = 0.6

# Default tasks directory (relative to repo root, or absolute)
TASKS_DIR = Path(__file__).parent.parent / "tasks"


def _load_task_def(task_id: str, tasks_dir: Path) -> dict:
    """Parse a task markdown file and return name/prompt/expected/grading sections."""
    md_path = tasks_dir / f"{task_id}.md"
    if not md_path.exists():
        return {}
    text = md_path.read_text(encoding="utf-8")

    # Strip YAML frontmatter
    if text.startswith("---"):
        end = text.find("\n---", 3)
        text = text[end + 4:] if end != -1 else text

    def _section(heading: str) -> str:
        pattern = rf"^##\s+{re.escape(heading)}\s*\n(.*?)(?=^##\s|\Z)"
        m = re.search(pattern, text, re.MULTILINE | re.DOTALL)
        return m.group(1).strip() if m else ""

    return {
        "prompt":            _section("Prompt"),
        "expected_behavior": _section("Expected Behavior"),
        "grading_criteria":  _section("Grading Criteria"),
    }

def _load_all_transcripts(results_dir: Path, run_id: str) -> dict[str, str]:
    """Parse {run_id}_*_transcripts.jsonl and return {task_id: last_assistant_text}."""
    merged = list(results_dir.glob(f"{run_id}_*_transcripts.jsonl"))
    if not merged:
        return {}

    task_map: dict[str, str] = {}
    current_task_id = ""
    with open(merged[0], encoding="utf-8") as f:
        for line in f:
            try:
                entry = json.loads(line)
            except json.JSONDecodeError:
                continue
            if entry.get("type") == "session":
                # session id format: task_XX_name_TIMESTAMP  →  strip trailing _DIGITS
                sid = entry.get("id", "")
                m = re.match(r"^(.*?)_(\d+)$", sid)
                current_task_id = m.group(1) if m else sid
            elif entry.get("type") == "message" and current_task_id:
                msg = entry.get("message", {})
                if msg.get("role") == "assistant":
                    for block in msg.get("content", []):
                        if block.get("type") == "text":
                            task_map[current_task_id] = block["text"]
    return task_map


COLOR_PASS   = "C6EFCE"   # green
COLOR_FAIL   = "FFC7CE"   # red
COLOR_HEADER = "4472C4"   # blue
COLOR_SECTION= "D9E1F2"   # light blue section title
COLOR_ALT    = "F2F2F2"   # alternating row grey


def _ts(ts: float) -> str:
    try:
        return datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return str(ts)


def _pct(v) -> str:
    try:
        return f"{float(v):.1%}"
    except Exception:
        return ""


def _status(task: dict) -> str:
    if task.get("timed_out") or task.get("status") != "success":
        return "FAIL"
    return "PASS"


def _failure_reason(task: dict) -> str:
    if task.get("timed_out"):
        return "执行超时"
    if task.get("status") != "success":
        return f"执行失败: {task.get('status', 'unknown')}"
    runs = task.get("grading", {}).get("runs", [])
    if not runs:
        return "无评分结果"
    run = runs[0]
    score = run.get("score", 0.0)
    notes = run.get("notes", "").strip()
    breakdown = run.get("breakdown", {})
    if score == 0.0 and not breakdown:
        return "Judge 解析失败（未返回 breakdown）"
    if score < PASS_THRESHOLD:
        if notes:
            return notes
        low = [f"{k}={v:.2f}" for k, v in breakdown.items()
               if isinstance(v, (int, float)) and v < PASS_THRESHOLD]
        return ("评分偏低: " + ", ".join(low)) if low else f"分数 {score:.0%} 未达阈值"
    return ""


def _header_style(cell, col_width=None):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _section_style(cell):
    cell.font = Font(bold=True, color="1F3864")
    cell.fill = PatternFill("solid", fgColor=COLOR_SECTION)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _kv(ws, row, key, value, bold_val=False):
    ws.cell(row=row, column=1, value=key).font = Font(bold=True)
    c = ws.cell(row=row, column=2, value=value)
    if bold_val:
        c.font = Font(bold=True, size=11)
    return row + 1


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 1: Task Results
# ─────────────────────────────────────────────────────────────────────────────
def _sheet_task_results(wb: openpyxl.Workbook, tasks: list,
                         tasks_dir: Path | None = None,
                         transcript_map: dict | None = None) -> None:
    ws = wb.active
    ws.title = "Task Results"

    columns = [
        ("task_ID",           22),
        ("task_name",              22),
        ("task_category",          14),
        ("grading_type",      13),
        ("prompt", 50),
        ("expected_behavior", 50),
        ("grading_criteria", 50),
        ("model_final_response", 60),
        ("exec_status",       11),
        ("is_timed_Out",          9),
        ("score",              8),
        ("score_breakdown",   40),
        ("judge_notes",       40),
        ("exec_time (s)",     12),
        ("input_tokens",      12),
        ("output_tokens",     12),
        ("requests",           9),

    ]

    for col_idx, (name, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        _header_style(cell)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    for i, task in enumerate(tasks, 2):
        grading  = task.get("grading", {})
        runs     = grading.get("runs", [])
        run0     = runs[0] if runs else {}
        usage    = task.get("usage", {})
        fm       = task.get("frontmatter", {})
        score    = grading.get("mean", 0.0)
        breakdown= run0.get("breakdown", {})
        bd_str   = ", ".join(f"{k}: {v:.2f}" for k, v in breakdown.items()) if breakdown else ""

        td = _load_task_def(task.get("task_id", ""), tasks_dir) if tasks_dir else {}
        model_resp = (transcript_map or {}).get(task.get("task_id", ""), "")

        row_data = [
            task.get("task_id", ""),
            fm.get("name", ""),
            fm.get("category", ""),
            run0.get("grading_type", ""),
            td.get("prompt", ""),
            td.get("expected_behavior", ""),
            td.get("grading_criteria", ""),
            model_resp,
            task.get("status", ""),
            "是" if task.get("timed_out") else "否",
            _pct(score),
            bd_str,
            run0.get("notes", ""),
            round(task.get("execution_time", 0), 1),
            usage.get("input_tokens", ""),
            usage.get("output_tokens", ""),
            usage.get("request_count", ""),
        ]

        alt = PatternFill("solid", fgColor=COLOR_ALT)

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if i % 2 == 0:
                cell.fill = alt


# ─────────────────────────────────────────────────────────────────────────────
# Sheet 2: Metrics Report
# ─────────────────────────────────────────────────────────────────────────────
def _sheet_metrics(wb: openpyxl.Workbook, result: dict, tasks: list) -> None:
    ws = wb.create_sheet("Metrics")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14

    efficiency = result.get("efficiency", {})
    passed  = sum(1 for t in tasks if _status(t) == "PASS")
    failed  = len(tasks) - passed
    overall = (sum(t.get("grading", {}).get("mean", 0.0) for t in tasks) / len(tasks)
               if tasks else 0.0)

    r = 1

    # ── Section 1: Overview ───────────────────────────────────────────────────
    c = ws.cell(row=r, column=1, value="▌ 总体概览")
    _section_style(c)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1

    r = _kv(ws, r, "模型 (Model)",       result.get("model", ""))
    r = _kv(ws, r, "Run ID",             result.get("run_id", ""))
    r = _kv(ws, r, "执行时间",            _ts(result.get("timestamp", 0)))
    r = _kv(ws, r, "Suite",              result.get("suite", ""))
    r = _kv(ws, r, "阈值 (Pass ≥)",      _pct(PASS_THRESHOLD))
    r += 1
    r = _kv(ws, r, "总任务数",            len(tasks))
    r = _kv(ws, r, "通过",               passed)
    r = _kv(ws, r, "失败",               failed)
    r = _kv(ws, r, "通过率",             _pct(passed / len(tasks)) if tasks else "")
    r = _kv(ws, r, "Overall Score",      _pct(overall), bold_val=True)
    r += 1

    # ── Section 2: Category Breakdown ────────────────────────────────────────
    c = ws.cell(row=r, column=1, value="▌ 分类得分")
    _section_style(c)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1

    cat_headers = ["Category", "任务数", "通过", "失败", "平均分", "通过率"]
    for ci, h in enumerate(cat_headers, 1):
        cell = ws.cell(row=r, column=ci, value=h)
        _header_style(cell)
    r += 1

    cat_data: dict[str, list] = defaultdict(list)
    for task in tasks:
        cat = str(task.get("frontmatter", {}).get("category", "unknown")).lower()
        cat_data[cat].append(task)

    for ci, (cat, cat_tasks) in enumerate(sorted(cat_data.items())):
        cat_passed = sum(1 for t in cat_tasks if _status(t) == "PASS")
        cat_avg    = sum(t.get("grading", {}).get("mean", 0.0) for t in cat_tasks) / len(cat_tasks)
        row_vals   = [
            cat,
            len(cat_tasks),
            cat_passed,
            len(cat_tasks) - cat_passed,
            _pct(cat_avg),
            _pct(cat_passed / len(cat_tasks)),
        ]
        fill = PatternFill("solid", fgColor=(COLOR_ALT if ci % 2 == 0 else "FFFFFF"))
        for col_idx, v in enumerate(row_vals, 1):
            cell = ws.cell(row=r, column=col_idx, value=v)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = fill
        r += 1
    r += 1

    # ── Section 3: Efficiency ─────────────────────────────────────────────────
    c = ws.cell(row=r, column=1, value="▌ 效率指标")
    _section_style(c)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1

    r = _kv(ws, r, "总 Token 数",              efficiency.get("total_tokens", ""))
    r = _kv(ws, r, "  Input Tokens",           efficiency.get("total_input_tokens", ""))
    r = _kv(ws, r, "  Output Tokens",          efficiency.get("total_output_tokens", ""))
    r = _kv(ws, r, "总 API 请求数",             efficiency.get("total_requests", ""))
    r = _kv(ws, r, "总执行时长 (s)",            efficiency.get("total_execution_time_seconds", ""))
    r = _kv(ws, r, "总费用 (USD)",             efficiency.get("total_cost_usd", ""))
    r = _kv(ws, r, "每任务平均 Token",          efficiency.get("tokens_per_task", ""))
    r = _kv(ws, r, "每千 Token 得分",
            round(efficiency["score_per_1k_tokens"], 4)
            if efficiency.get("score_per_1k_tokens") else "N/A")
    r = _kv(ws, r, "每美元得分",
            round(efficiency["score_per_dollar"], 2)
            if efficiency.get("score_per_dollar") else "N/A")
    r += 1

    # ── Section 4: Per-task efficiency ───────────────────────────────────────
    c = ws.cell(row=r, column=1, value="▌ 各任务效率明细")
    _section_style(c)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1

    pt_headers = ["Task ID", "得分", "Total Tokens", "Cost (USD)", "Pass/Fail", "Token/分"]
    for ci, h in enumerate(pt_headers, 1):
        cell = ws.cell(row=r, column=ci, value=h)
        _header_style(cell)
    r += 1

    per_task = efficiency.get("per_task", [])
    task_map = {t.get("task_id"): t for t in tasks}
    for ci, pt in enumerate(per_task):
        tid    = pt.get("task_id", "")
        status = _status(task_map[tid]) if tid in task_map else ""
        tps    = pt.get("tokens_per_score_point")
        row_vals = [
            tid,
            _pct(pt.get("score", 0)),
            pt.get("total_tokens", ""),
            pt.get("cost_usd", ""),
            status,
            round(tps, 1) if tps else "N/A",
        ]
        fill = PatternFill("solid", fgColor=(COLOR_PASS if status == "PASS" else
                           COLOR_FAIL if status == "FAIL" else
                           COLOR_ALT if ci % 2 == 0 else "FFFFFF"))
        for col_idx, v in enumerate(row_vals, 1):
            cell = ws.cell(row=r, column=col_idx, value=v)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = fill
        r += 1


# ─────────────────────────────────────────────────────────────────────────────
def build_report(result: dict, wb: openpyxl.Workbook,
                 tasks_dir: Path | None = None,
                 results_dir: Path | None = None) -> None:
    tasks  = result.get("tasks", [])
    run_id = result.get("run_id", "")
    transcript_map = _load_all_transcripts(results_dir, run_id) if results_dir else {}
    _sheet_task_results(wb, tasks, tasks_dir=tasks_dir, transcript_map=transcript_map)
    _sheet_metrics(wb, result, tasks)


def _export_one(json_path: Path, out_path: Path) -> None:
    with open(json_path, encoding="utf-8") as f:
        result = json.load(f)

    tasks_dir   = TASKS_DIR if TASKS_DIR.is_dir() else None
    results_dir = json_path.parent

    wb = openpyxl.Workbook()
    build_report(result, wb, tasks_dir=tasks_dir, results_dir=results_dir)
    wb.save(out_path)

    tasks   = result.get("tasks", [])
    passed  = sum(1 for t in tasks if _status(t) == "PASS")
    overall = sum(t.get("grading", {}).get("mean", 0.0) for t in tasks) / len(tasks) if tasks else 0.0
    print(f"Saved: {out_path}")
    print(f"  {len(tasks)} tasks | {passed} passed | {len(tasks)-passed} failed | Overall {overall:.1%}")


def main() -> None:
    global PASS_THRESHOLD  # noqa: PLW0603
    parser = argparse.ArgumentParser(description="Export benchmark JSON to Excel")
    parser.add_argument("input", help="Result JSON file or directory")
    parser.add_argument("-o", "--output", help="Output .xlsx path (default: same name as input)")
    parser.add_argument("--threshold", type=float, default=PASS_THRESHOLD,
                        help=f"Pass threshold 0~1 (default: {PASS_THRESHOLD})")
    args = parser.parse_args()
    PASS_THRESHOLD = args.threshold

    input_path = Path(args.input)
    if input_path.is_dir():
        json_files = sorted(input_path.glob("*.json"))
        if not json_files:
            print(f"No JSON files found in {input_path}")
            sys.exit(1)
        for jf in json_files:
            _export_one(jf, jf.with_suffix(".xlsx"))
    else:
        out = Path(args.output) if args.output else input_path.with_suffix(".xlsx")
        _export_one(input_path, out)


if __name__ == "__main__":
    main()